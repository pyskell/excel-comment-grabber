#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
from openpyxl import load_workbook

try:
  file_name = str(sys.argv[1])
except IndexError:
  exit("Usage: " + sys.argv[0] + " filename.xlsx")

workbook = load_workbook(file_name)

sheet_names = workbook.get_sheet_names()

for sheet in sheet_names:
  current_sheet = workbook.get_sheet_by_name(sheet)
  sheet_comments = []

  for row in current_sheet.iter_rows():
    for cell in row:

      comment = cell.comment
      if comment:

        comment_lines = comment.text.split('\n')[1:]
        for line in comment_lines:
          sheet_comments.append(line)

  if sheet_comments:
    print(current_sheet['A3'].value,":", " ".join(str(x) for x in sheet_comments))